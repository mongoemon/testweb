#!/usr/bin/env python3
"""
ShoesHub Test Case Excel Generator
Run: python create_test_excel.py
Output: ShoesHub_Test_Cases.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date

TODAY = date.today().isoformat()

# ─── Helpers ─────────────────────────────────────────────────────────────────
def fx(hex_color): return PatternFill("solid", fgColor=hex_color)
def ft(bold=False, color="1F2937", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def al(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bd(color="D1D5DB", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, val, bg="1F2937", fc="FFFFFF", size=10, h="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fx(bg); c.font = ft(True, fc, size); c.alignment = al(h, "center", True)
    c.border = bd("374151", "medium"); return c

def cell(ws, row, col, val, bg=None, bold=False, fc="1F2937", h="left", sz=9, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = fx(bg)
    c.font = ft(bold, fc, sz, italic); c.alignment = al(h, "top", True)
    c.border = bd(); return c

PRIORITY_STYLE = {
    "P0": ("DC2626","FFFFFF"),  # Red — Smoke/Critical
    "P1": ("EA580C","FFFFFF"),  # Orange — High
    "P2": ("CA8A04","FFFFFF"),  # Amber — Medium
    "P3": ("16A34A","FFFFFF"),  # Green — Low
}
AUTO_STYLE = {
    "Automated":       ("1D4ED8","FFFFFF"),
    "To Be Automated": ("7C3AED","FFFFFF"),
    "In Progress":     ("0369A1","FFFFFF"),
    "Manual Only":     ("6B7280","FFFFFF"),
}

# ─── Test Case Data ───────────────────────────────────────────────────────────
# Columns: id, module, feature, name_th, priority, type_, role,
#          precondition, steps, expected, testids, api, data_ref,
#          auto_status, script_path, tags

TC = [
  # ── AUTH ──
  ("TC-AUTH-01","Authentication","Register","สมัครสมาชิกสำเร็จ","P1","Functional","Guest",
   "ยังไม่มี account นี้ในระบบ",
   "1. เปิด /register.html\n2. กรอก Full Name: 'Test New User'\n3. Username: 'newuser01'\n4. Email: 'newuser01@test.com'\n5. Password: 'pass1234'\n6. Confirm: 'pass1234'\n7. คลิก Submit",
   "- Toast success ปรากฏ\n- Redirect ไป /\n- Navbar แสดงชื่อ 'newuser01'\n- ไม่มีปุ่ม Login/Register",
   "nav-register, toast[data-type=success], nav-user-menu",
   "POST /api/auth/register","TD-AUTH-01","To Be Automated","tests/auth/register.spec.ts","auth,register,happy-path"),

  ("TC-AUTH-02","Authentication","Register","สมัคร — username ซ้ำ","P1","Negative","Guest",
   "มี 'testuser' อยู่ในระบบแล้ว",
   "1. เปิด /register.html\n2. Username: 'testuser' (ซ้ำ)\n3. กรอกข้อมูลอื่นให้ครบ\n4. คลิก Submit",
   "- ไม่ redirect\n- แสดง error 'Username already taken'\n- ยังอยู่หน้า register",
   "nav-register","POST /api/auth/register","TD-AUTH-02","To Be Automated","tests/auth/register.spec.ts","auth,register,negative"),

  ("TC-AUTH-03","Authentication","Register","สมัคร — password สั้น (< 6 ตัว)","P2","Negative","Guest",
   "—",
   "1. เปิด /register.html\n2. Password: 'abc'\n3. คลิก Submit",
   "- แสดง error validation\n- ไม่สร้าง account",
   "","POST /api/auth/register","","To Be Automated","tests/auth/register.spec.ts","auth,register,negative,validation"),

  ("TC-AUTH-04","Authentication","Register","สมัคร — password ไม่ตรงกัน","P2","Negative","Guest",
   "—",
   "1. เปิด /register.html\n2. Password: 'pass1234'\n3. Confirm: 'pass9999'\n4. คลิก Submit",
   "- Error 'รหัสผ่านไม่ตรงกัน'\n- ไม่ส่ง API request",
   "","—","","To Be Automated","tests/auth/register.spec.ts","auth,register,negative,validation"),

  ("TC-AUTH-05","Authentication","Login","Login สำเร็จ — User","P0","Smoke","Guest",
   "มี testuser / test1234",
   "1. เปิด /login.html\n2. Username: 'testuser'\n3. Password: 'test1234'\n4. คลิก Login",
   "- Redirect ไป /\n- Navbar แสดง 'testuser' + cart icon\n- ไม่เห็น Login/Register",
   "nav-login, nav-user-menu, nav-cart","POST /api/auth/login","TD-AUTH-05","To Be Automated","tests/auth/login.spec.ts","auth,login,smoke,happy-path"),

  ("TC-AUTH-06","Authentication","Login","Login สำเร็จ — Admin","P0","Smoke","Guest",
   "มี admin / admin1234",
   "1. Login ด้วย admin / admin1234",
   "- Navbar แสดง nav-admin link\n- เห็น cart, orders, admin links",
   "nav-admin","POST /api/auth/login","TD-AUTH-06","To Be Automated","tests/auth/login.spec.ts","auth,login,smoke,admin"),

  ("TC-AUTH-07","Authentication","Login","Login ผิด — wrong password","P1","Negative","Guest",
   "—",
   "1. เปิด /login.html\n2. Username: 'testuser'\n3. Password: 'wrongpass'\n4. คลิก Login",
   "- ไม่ redirect\n- แสดง error message ใน form\n- ยังอยู่หน้า login",
   "","POST /api/auth/login","TD-AUTH-07","To Be Automated","tests/auth/login.spec.ts","auth,login,negative"),

  ("TC-AUTH-08","Authentication","Logout","Logout สำเร็จ","P1","Functional","User",
   "Login แล้ว",
   "1. คลิก user menu (ชื่อ user)\n2. คลิก 'ออกจากระบบ'",
   "- Redirect ไป /\n- Navbar แสดง Login/Register\n- ไม่เห็น cart/orders",
   "nav-user-menu, nav-logout, nav-login","—","","To Be Automated","tests/auth/logout.spec.ts","auth,logout"),

  ("TC-AUTH-09","Authentication","Access Control","เข้าหน้า protected ขณะ logout","P1","Functional","Guest",
   "ไม่ได้ login",
   "1. เปิด /cart.html โดยตรง",
   "- Redirect ไป /login.html?redirect=%2Fcart.html\n- หลัง login → redirect กลับ /cart.html",
   "","—","","To Be Automated","tests/auth/access-control.spec.ts","auth,security,redirect"),

  ("TC-AUTH-10","Authentication","Access Control","User เข้าหน้า Admin — redirect","P1","Security","User",
   "Login ด้วย testuser",
   "1. Login ด้วย testuser\n2. เปิด /admin/index.html",
   "- Redirect ไป /login.html\n- ไม่เห็น dashboard content",
   "","—","","To Be Automated","tests/auth/access-control.spec.ts","auth,security,admin"),

  # ── HOME ──
  ("TC-HOME-01","Homepage","Page Load","หน้า Home โหลดสำเร็จ","P0","Smoke","Guest",
   "ระบบทำงาน มี seed data",
   "1. เปิด http://localhost:8000/",
   "- Hero section แสดง\n- Categories ≥ 1 card\n- Featured products ≥ 1 รายการ\n- ไม่มี console error",
   "hero-section, categories-section, featured-products","—","","To Be Automated","tests/home/home.spec.ts","home,smoke"),

  ("TC-HOME-02","Homepage","Navigation","Hero CTA ไป Products","P2","UI","Guest",
   "—",
   "1. เปิดหน้า Home\n2. คลิก 'เลือกซื้อเลย →'",
   "- Navigate ไป /products.html\n- แสดงรายการสินค้า",
   "hero-shop-btn","—","","To Be Automated","tests/home/home.spec.ts","home,navigation"),

  ("TC-HOME-03","Homepage","Navigation","คลิก Category card","P2","UI","Guest",
   "—",
   "1. เปิดหน้า Home\n2. คลิก category card (เช่น Running)",
   "- Navigate ไป /products.html?category={id}\n- Products ถูก filter",
   "categories-section","—","","To Be Automated","tests/home/home.spec.ts","home,navigation,filter"),

  # ── PRODUCTS ──
  ("TC-PROD-01","Products","Listing","แสดงสินค้าทั้งหมด","P0","Smoke","Guest",
   "มี seed data 12 รายการ",
   "1. เปิด /products.html",
   "- Product cards ≥ 12 รายการ\n- count label แสดงตัวเลข\n- แต่ละ card มี: ชื่อ, ราคา, รูป",
   "products-grid, product-card, product-count","GET /api/products","","To Be Automated","tests/products/catalog.spec.ts","products,smoke,listing"),

  ("TC-PROD-02","Products","Search","ค้นหาด้วยชื่อสินค้า","P1","Functional","Guest",
   "—",
   "1. เปิด /products.html\n2. พิมพ์ 'Nike' ใน search box",
   "- แสดงเฉพาะสินค้า Nike\n- count อัปเดต",
   "search-input, product-card, product-count","GET /api/products?search=Nike","TD-PROD-02","To Be Automated","tests/products/search.spec.ts","products,search"),

  ("TC-PROD-03","Products","Search","ค้นหาแล้วไม่พบสินค้า","P2","Negative","Guest",
   "—",
   "1. พิมพ์ 'xxxxnotexist' ใน search box",
   "- แสดง 'ไม่พบสินค้า'\n- ไม่มี product card",
   "search-input","GET /api/products?search=xxxxnotexist","","To Be Automated","tests/products/search.spec.ts","products,search,negative"),

  ("TC-PROD-04","Products","Filter","Filter ตาม Category","P1","Functional","Guest",
   "—",
   "1. เลือก category 'Running' จาก dropdown",
   "- แสดงเฉพาะสินค้า Running\n- count อัปเดต",
   "category-select, product-card","GET /api/products?category_id=1","","To Be Automated","tests/products/filter.spec.ts","products,filter,category"),

  ("TC-PROD-05","Products","Filter","Filter ตาม Brand","P2","Functional","Guest",
   "—",
   "1. เลือก brand 'Nike'",
   "- แสดงเฉพาะสินค้า Nike brand\n- count อัปเดต",
   "brand-select","GET /api/products?brand=Nike","","To Be Automated","tests/products/filter.spec.ts","products,filter,brand"),

  ("TC-PROD-06","Products","Sort","Sort ราคา ต่ำ→สูง","P2","Functional","Guest",
   "—",
   "1. เลือก sort 'ราคา: ต่ำ→สูง'",
   "- ราคา card แรก ≤ card ถัดไปทุกใบ",
   "sort-select, product-card-price","GET /api/products?sort=price_asc","","To Be Automated","tests/products/sort.spec.ts","products,sort"),

  ("TC-PROD-07","Products","Sort","Sort ราคา สูง→ต่ำ","P2","Functional","Guest",
   "—",
   "1. เลือก sort 'ราคา: สูง→ต่ำ'",
   "- ราคา card แรก ≥ card ถัดไปทุกใบ",
   "sort-select, product-card-price","GET /api/products?sort=price_desc","","To Be Automated","tests/products/sort.spec.ts","products,sort"),

  ("TC-PROD-08","Products","Filter","Filter หลายเงื่อนไขพร้อมกัน","P2","Functional","Guest",
   "—",
   "1. Category: Casual\n2. Brand: Nike\n3. Size: 42",
   "- แสดงเฉพาะสินค้าที่ตรงทุกเงื่อนไข (AND)",
   "category-select, brand-select, size-select","GET /api/products?category_id=2&brand=Nike&size=42","","To Be Automated","tests/products/filter.spec.ts","products,filter,combined"),

  ("TC-PROD-09","Products","Navigation","คลิกไป Product Detail","P1","Functional","Guest",
   "—",
   "1. คลิก product card ใดก็ได้",
   "- Navigate ไป /product.html?id={id}\n- URL มี id ที่ถูกต้อง",
   "product-card","—","","To Be Automated","tests/products/catalog.spec.ts","products,navigation"),

  # ── PRODUCT DETAIL ──
  ("TC-DETAIL-01","Product Detail","Page Load","แสดงข้อมูลสินค้าครบ","P1","Functional","Guest",
   "สินค้า id=1 มีอยู่",
   "1. เปิด /product.html?id=1",
   "- แสดง: ชื่อ, ราคา, brand, รูป, คำอธิบาย\n- Size buttons แสดง (ถ้ามี)\n- Stock status แสดง",
   "product-name, product-price, product-brand, product-stock, size-buttons","GET /api/products/1","","To Be Automated","tests/products/detail.spec.ts","product-detail,smoke"),

  ("TC-DETAIL-02","Product Detail","Access Control","Guest กด Add to Cart → redirect","P1","Functional","Guest",
   "ยังไม่ได้ login",
   "1. เปิด product detail\n2. เลือก size\n3. คลิก Add to Cart",
   "- Redirect ไป /login.html?redirect=...\n- ไม่เพิ่มลงตะกร้า",
   "add-to-cart-btn","—","","To Be Automated","tests/products/detail.spec.ts","product-detail,security,redirect"),

  ("TC-DETAIL-03","Product Detail","Validation","Add to Cart ไม่เลือก size","P1","Negative","User",
   "Login แล้ว, สินค้ามี sizes",
   "1. Login\n2. เปิด product detail ที่มี sizes\n3. ไม่เลือก size\n4. คลิก Add to Cart",
   "- Error 'กรุณาเลือกไซส์'\n- ไม่ส่ง request\n- cart count ไม่เพิ่ม",
   "size-error, add-to-cart-btn, cart-count","—","","To Be Automated","tests/products/detail.spec.ts","product-detail,negative,validation"),

  ("TC-DETAIL-04","Product Detail","Cart","Add to Cart สำเร็จ","P0","Smoke","User",
   "Login แล้ว, สินค้ามี stock",
   "1. เปิด product detail (id=1)\n2. เลือก size '42'\n3. Qty: 2\n4. คลิก Add to Cart",
   "- Toast 'เพิ่มลงตะกร้าแล้ว!'\n- cart badge อัปเดต",
   "size-btn-42, add-to-cart-btn, toast, cart-count","POST /api/cart","TD-DETAIL-04","To Be Automated","tests/products/detail.spec.ts","product-detail,cart,smoke"),

  ("TC-DETAIL-05","Product Detail","Stock","สินค้าหมดสต็อก","P2","Functional","Guest",
   "สินค้ามี stock=0",
   "1. เปิด product detail ของสินค้า stock=0",
   "- แสดง 'หมดสต็อก'\n- ปุ่ม Add to Cart disable หรือหาย",
   "product-stock, add-to-cart-btn","—","","To Be Automated","tests/products/detail.spec.ts","product-detail,stock"),

  ("TC-DETAIL-06","Product Detail","Stock","สินค้า stock น้อย (≤5)","P3","UI","Guest",
   "สินค้ามี 1 ≤ stock ≤ 5",
   "1. เปิด product detail ของสินค้า low stock",
   "- แสดง 'เหลือ {n} คู่' พร้อม warning style",
   "product-stock","—","","To Be Automated","tests/products/detail.spec.ts","product-detail,stock,ui"),

  # ── CART ──
  ("TC-CART-01","Cart","Empty State","ตะกร้าว่าง","P2","Functional","User",
   "Login แล้ว, ตะกร้าว่าง",
   "1. เปิด /cart.html",
   "- Empty state 'ตะกร้าของคุณว่างเปล่า'\n- ปุ่ม 'เลือกซื้อสินค้า' ลิงก์ไป /products.html",
   "empty-cart","GET /api/cart","","To Be Automated","tests/cart/cart.spec.ts","cart,empty-state"),

  ("TC-CART-02","Cart","View","ตะกร้ามีสินค้า","P1","Functional","User",
   "มีสินค้า ≥ 1 ในตะกร้า",
   "1. เปิด /cart.html",
   "- แสดงรายการสินค้าครบ\n- ชื่อ, size, qty, subtotal ถูกต้อง\n- Total แสดง",
   "cart-items, cart-item, item-name, cart-total, checkout-btn","GET /api/cart","","To Be Automated","tests/cart/cart.spec.ts","cart,view"),

  ("TC-CART-03","Cart","Update","แก้ไขจำนวนสินค้า","P1","Functional","User",
   "มีสินค้าในตะกร้า",
   "1. เปิด /cart.html\n2. เปลี่ยน qty ของ item แรกเป็น 3",
   "- subtotal = price × 3\n- total อัปเดต",
   "item-quantity, item-subtotal, cart-total","PUT /api/cart/{id}","","To Be Automated","tests/cart/cart.spec.ts","cart,update"),

  ("TC-CART-04","Cart","Remove","ลบสินค้าออกจากตะกร้า","P1","Functional","User",
   "มีสินค้า ≥ 2 ในตะกร้า",
   "1. เปิด /cart.html\n2. คลิก × ของ item แรก",
   "- Item หายออกจาก list\n- Total อัปเดต\n- Toast 'ลบสินค้าแล้ว'",
   "remove-item-{id}, cart-items, toast","DELETE /api/cart/{id}","","To Be Automated","tests/cart/cart.spec.ts","cart,remove"),

  ("TC-CART-05","Cart","Clear","ล้างตะกร้าทั้งหมด","P2","Functional","User",
   "มีสินค้าในตะกร้า",
   "1. คลิก '🗑 ล้างตะกร้า'\n2. กด OK ใน confirm dialog",
   "- Cart ว่าง แสดง empty state\n- cart badge = 0 หรือซ่อน",
   "clear-cart-btn, empty-cart, cart-count","DELETE /api/cart/clear","","To Be Automated","tests/cart/cart.spec.ts","cart,clear"),

  ("TC-CART-06","Cart","Badge","Cart badge อัปเดตใน navbar","P2","UI","User",
   "Login แล้ว",
   "1. เพิ่มสินค้า 3 ชิ้น\n2. Navigate ไปหน้าอื่น",
   "- Badge แสดงตัวเลขที่ถูกต้อง",
   "cart-count","GET /api/cart","","To Be Automated","tests/cart/cart.spec.ts","cart,badge,ui"),

  # ── CHECKOUT ──
  ("TC-CHK-01","Checkout","Page Load","Checkout form โหลดพร้อม summary","P1","Functional","User",
   "มีสินค้าในตะกร้า",
   "1. เปิด /checkout.html",
   "- Form จัดส่งแสดง\n- Order summary ด้านขวาแสดงรายการและ total",
   "checkout-form, order-summary, checkout-total","GET /api/cart","","To Be Automated","tests/checkout/checkout.spec.ts","checkout,page-load"),

  ("TC-CHK-02","Checkout","Redirect","ตะกร้าว่าง redirect ไป cart","P1","Functional","User",
   "ตะกร้าว่าง",
   "1. เปิด /checkout.html โดยตรง",
   "- Redirect ไป /cart.html อัตโนมัติ",
   "","GET /api/cart","","To Be Automated","tests/checkout/checkout.spec.ts","checkout,redirect"),

  ("TC-CHK-03","Checkout","Pre-fill","Pre-fill ที่อยู่จาก profile","P2","Functional","User",
   "บันทึก default address ใน profile แล้ว",
   "1. บันทึกที่อยู่ใน /profile.html\n2. เพิ่มสินค้าลงตะกร้า\n3. เปิด /checkout.html",
   "- Fields shipping-name, address, city, postal, phone ถูก pre-fill\n- Payment method radio pre-select",
   "shipping-name, shipping-address, shipping-city","GET /api/auth/me","","To Be Automated","tests/checkout/checkout.spec.ts","checkout,pre-fill,e2e"),

  ("TC-CHK-04","Checkout","Order","สั่งซื้อสำเร็จ","P0","Smoke","User",
   "มีสินค้าในตะกร้า",
   "1. กรอก Shipping Name: 'John Doe'\n2. Address: '123 Test Rd'\n3. City: 'Bangkok'\n4. Postal: '10110'\n5. Phone: '081-000-0000'\n6. Payment: Credit Card\n7. คลิก 'ยืนยันคำสั่งซื้อ'",
   "- Redirect ไป /order-detail.html?id={id}&success=1\n- Success banner แสดง\n- Cart cleared (badge=0)",
   "place-order-btn, success-banner, cart-count","POST /api/orders","TD-CHK-04","To Be Automated","tests/checkout/checkout.spec.ts","checkout,order,smoke,e2e"),

  ("TC-CHK-05","Checkout","Validation","Validation — field ว่าง","P1","Negative","User",
   "—",
   "1. ปล่อย Shipping Name ว่าง\n2. คลิก 'ยืนยันคำสั่งซื้อ'",
   "- Error 'กรุณากรอกข้อมูลให้ครบถ้วน'\n- ไม่ส่ง request\n- focus ไปที่ field ว่าง",
   "form-error, place-order-btn","—","","To Be Automated","tests/checkout/checkout.spec.ts","checkout,negative,validation"),

  # ── ORDERS ──
  ("TC-ORD-01","Orders","Empty State","ไม่มีออเดอร์ — empty state","P2","Functional","User",
   "Account ใหม่ ยังไม่เคยสั่ง",
   "1. เปิด /orders.html",
   "- Empty state 'ยังไม่มีคำสั่งซื้อ'\n- ปุ่ม 'เลือกซื้อสินค้า'",
   "empty-orders","GET /api/orders","","To Be Automated","tests/orders/orders.spec.ts","orders,empty-state"),

  ("TC-ORD-02","Orders","List","แสดงรายการออเดอร์","P1","Functional","User",
   "มีออเดอร์ ≥ 1",
   "1. เปิด /orders.html",
   "- Order cards ของ user นี้เท่านั้น\n- แต่ละ card: ID, วันที่, items, total, status",
   "orders-container, order-card, order-status","GET /api/orders","","To Be Automated","tests/orders/orders.spec.ts","orders,list"),

  ("TC-ORD-03","Orders","Detail","ดูรายละเอียด order","P1","Functional","User",
   "มีออเดอร์",
   "1. คลิก 'ดูรายละเอียด →' บน order card",
   "- Navigate ไป /order-detail.html?id={id}\n- แสดง: order#, status, items, ที่อยู่, payment, total",
   "order-number, order-status-badge, shipping-info, payment-info","GET /api/orders/{id}","","To Be Automated","tests/orders/orders.spec.ts","orders,detail"),

  ("TC-ORD-04","Orders","Security","User เข้า order ของคนอื่น","P1","Security","User",
   "Login ด้วย testuser",
   "1. เปิด /order-detail.html?id=9999 (ไม่ใช่ของตัวเอง)",
   "- แสดง 'ไม่พบคำสั่งซื้อ'\n- ไม่เห็นข้อมูล order ของคนอื่น",
   "","GET /api/orders/9999","","To Be Automated","tests/orders/orders.spec.ts","orders,security"),

  # ── PROFILE ──
  ("TC-PROF-01","Profile","Page Load","โหลดหน้า profile พร้อมข้อมูล","P1","Functional","User",
   "Login แล้ว",
   "1. เปิด /profile.html",
   "- profile-email แสดง email ปัจจุบัน\n- profile-fullname แสดง full name\n- Form password ว่าง",
   "profile-email, profile-fullname, form-personal","GET /api/auth/me","","To Be Automated","tests/profile/profile.spec.ts","profile,page-load"),

  ("TC-PROF-02","Profile","Personal Info","แก้ไขชื่อและ email สำเร็จ","P1","Functional","User",
   "Login แล้ว",
   "1. แก้ Full Name: 'Updated Name'\n2. Email: 'updated@test.com'\n3. คลิก 'บันทึกข้อมูล'",
   "- Toast 'บันทึกข้อมูลส่วนตัวแล้ว'\n- Reload → fields แสดงค่าใหม่",
   "profile-fullname, profile-email, profile-personal-save, toast","PUT /api/auth/profile","TD-PROF-02","To Be Automated","tests/profile/profile.spec.ts","profile,personal-info"),

  ("TC-PROF-03","Profile","Password","เปลี่ยนรหัสผ่านสำเร็จ","P1","Functional","User",
   "Login ด้วย testuser / test1234",
   "1. Current: 'test1234'\n2. New: 'newpass99'\n3. Confirm: 'newpass99'\n4. คลิก 'เปลี่ยนรหัสผ่าน'",
   "- Toast 'เปลี่ยนรหัสผ่านแล้ว'\n- Form reset เป็นว่าง\n- Login ด้วย password ใหม่ได้",
   "profile-current-password, profile-new-password, profile-password-save, toast","PUT /api/auth/password","TD-PROF-03","To Be Automated","tests/profile/profile.spec.ts","profile,password"),

  ("TC-PROF-04","Profile","Password","เปลี่ยนรหัสผ่าน — current ผิด","P1","Negative","User",
   "—",
   "1. Current: 'wrongpassword'\n2. New + Confirm ถูกต้อง\n3. Submit",
   "- Error 'Current password is incorrect'\n- Password ไม่เปลี่ยน",
   "profile-password-error","PUT /api/auth/password","","To Be Automated","tests/profile/profile.spec.ts","profile,password,negative"),

  ("TC-PROF-05","Profile","Password","เปลี่ยนรหัสผ่าน — new ไม่ตรงกัน","P1","Negative","User",
   "—",
   "1. Current ถูกต้อง\n2. New: 'pass1111'\n3. Confirm: 'pass2222'\n4. Submit",
   "- Error 'รหัสผ่านใหม่ไม่ตรงกัน'\n- ไม่ส่ง request",
   "profile-password-error","—","","To Be Automated","tests/profile/profile.spec.ts","profile,password,negative"),

  ("TC-PROF-06","Profile","Address","บันทึกที่อยู่จัดส่งเริ่มต้น","P2","Functional","User",
   "Login แล้ว",
   "1. Shipping Name: 'Test Shipper'\n2. Address: '456 Sample St'\n3. City: 'Chiang Mai'\n4. Postal: '50000'\n5. Phone: '089-000-0000'\n6. Payment: Bank Transfer\n7. คลิก 'บันทึกที่อยู่'",
   "- Toast 'บันทึกที่อยู่แล้ว'\n- Reload → fields แสดงค่าที่บันทึก\n- Radio bank_transfer checked",
   "profile-shipping-name, profile-address-save, toast, payment-bank-transfer","PUT /api/auth/address","TD-PROF-06","To Be Automated","tests/profile/profile.spec.ts","profile,address"),

  ("TC-PROF-07","Profile","E2E","ที่อยู่ที่บันทึกไว้ auto-fill ใน Checkout","P2","E2E","User",
   "บันทึก default address แล้ว (TC-PROF-06)",
   "1. เพิ่มสินค้าลงตะกร้า\n2. เปิด /checkout.html",
   "- shipping-name = 'Test Shipper'\n- shipping-city = 'Chiang Mai'\n- payment-bank-transfer checked",
   "shipping-name, shipping-city, payment-bank-transfer","GET /api/auth/me","","To Be Automated","tests/profile/profile.spec.ts","profile,checkout,e2e"),

  # ── ADMIN DASHBOARD ──
  ("TC-DASH-01","Admin Dashboard","Stats","Stats cards โหลดครบ","P1","Functional","Admin",
   "Login ด้วย admin, มี seed data",
   "1. เปิด /admin/index.html",
   "- stat-total-orders แสดงตัวเลข\n- stat-pending-orders แสดงตัวเลข\n- stat-revenue แสดง ฿\n- stat-products แสดง\n- stat-users แสดง",
   "stats-grid, stat-total-orders, stat-revenue, stat-products, stat-users","GET /api/admin/stats","","To Be Automated","tests/admin/dashboard.spec.ts","admin,dashboard,stats"),

  ("TC-DASH-02","Admin Dashboard","Recent Orders","Recent orders table","P2","Functional","Admin",
   "มีออเดอร์ ≥ 1",
   "1. เปิด /admin/index.html",
   "- recent-orders แสดง table\n- rows มี: order link, username, วันที่, total, status",
   "recent-orders","GET /api/admin/stats","","To Be Automated","tests/admin/dashboard.spec.ts","admin,dashboard,orders"),

  ("TC-DASH-03","Admin Dashboard","Access Control","Non-admin เข้า dashboard — redirect","P1","Security","User",
   "Login ด้วย testuser",
   "1. เปิด /admin/index.html",
   "- Redirect ไป /login.html\n- ไม่เห็น dashboard",
   "","—","","To Be Automated","tests/admin/dashboard.spec.ts","admin,security,redirect"),

  # ── ADMIN PRODUCTS ──
  ("TC-APROD-01","Admin Products","Listing","รายการสินค้า","P1","Functional","Admin",
   "Login ด้วย admin",
   "1. เปิด /admin/products.html",
   "- product-table แสดง ≥ 12 rows\n- แต่ละ row มี: รูป, ชื่อ, brand, category, ราคา, stock, ปุ่ม edit/delete",
   "product-table, product-row","GET /api/products","","To Be Automated","tests/admin/products.spec.ts","admin,products,listing"),

  ("TC-APROD-02","Admin Products","Create","เพิ่มสินค้าใหม่สำเร็จ","P1","Functional","Admin",
   "—",
   "1. คลิก '+ เพิ่มสินค้า'\n2. Name: 'Test Shoe X'\n3. Brand: 'TestBrand'\n4. Category: Running\n5. Price: 1990\n6. Stock: 10\n7. Sizes: '40,41,42'\n8. คลิก 'บันทึก'",
   "- Modal ปิด\n- Toast 'เพิ่มสินค้าแล้ว'\n- Table แสดง 'Test Shoe X'",
   "add-product-btn, product-modal, product-form-name, product-save-btn, toast","POST /api/products","TD-APROD-02","To Be Automated","tests/admin/products.spec.ts","admin,products,create"),

  ("TC-APROD-03","Admin Products","Validation","เพิ่มสินค้า — ไม่กรอก required","P1","Negative","Admin",
   "—",
   "1. คลิก '+ เพิ่มสินค้า'\n2. ปล่อย Name ว่าง\n3. Price: 1990\n4. คลิก 'บันทึก'",
   "- Error 'กรุณากรอกชื่อสินค้าและราคา'\n- Modal ไม่ปิด",
   "product-form-error","—","","To Be Automated","tests/admin/products.spec.ts","admin,products,negative,validation"),

  ("TC-APROD-04","Admin Products","Update","แก้ไขสินค้า","P1","Functional","Admin",
   "มีสินค้าใน list",
   "1. คลิก Edit ของสินค้าแรก\n2. แก้ Stock: 99\n3. คลิก 'บันทึก'",
   "- Modal ปิด\n- Toast 'อัปเดตสินค้าแล้ว'\n- Row แสดง stock=99",
   "edit-product-{id}, product-form-stock, product-save-btn, toast","PUT /api/products/{id}","","To Be Automated","tests/admin/products.spec.ts","admin,products,update"),

  ("TC-APROD-05","Admin Products","UI","Modal ปิดด้วยปุ่ม X","P3","UI","Admin",
   "—",
   "1. คลิก '+ เพิ่มสินค้า'\n2. คลิกปุ่ม X มุมบนขวา",
   "- Modal ซ่อน\n- ไม่บันทึกข้อมูล",
   "product-modal, modal-close-btn","—","","To Be Automated","tests/admin/products.spec.ts","admin,products,ui,modal"),

  ("TC-APROD-06","Admin Products","UI","Modal ปิดด้วยคลิก backdrop","P3","UI","Admin",
   "—",
   "1. เปิด modal\n2. คลิก overlay สีดำนอก modal",
   "- Modal ปิด",
   "product-modal","—","","Manual Only","tests/admin/products.spec.ts","admin,products,ui,modal"),

  ("TC-APROD-07","Admin Products","Delete","ลบสินค้า","P2","Functional","Admin",
   "มีสินค้าที่ไม่ผูกกับ active order",
   "1. คลิก Delete ของ 'Test Shoe X'\n2. กด OK ใน confirm",
   "- Toast 'ลบสินค้าแล้ว'\n- สินค้านั้นหายจาก list",
   "delete-product-{id}, toast","DELETE /api/products/{id}","","To Be Automated","tests/admin/products.spec.ts","admin,products,delete"),

  # ── ADMIN ORDERS ──
  ("TC-AORD-01","Admin Orders","Listing","แสดง orders ทั้งหมด","P1","Functional","Admin",
   "มี order ≥ 1",
   "1. เปิด /admin/orders.html",
   "- orders-table แสดง rows ของทุก user\n- แต่ละ row: order link, username, วันที่, ที่อยู่, total, status, dropdown",
   "orders-table, order-row","GET /api/admin/orders","","To Be Automated","tests/admin/orders.spec.ts","admin,orders,listing"),

  ("TC-AORD-02","Admin Orders","Filter","Filter order ตาม status","P1","Functional","Admin",
   "—",
   "1. เลือก 'รอดำเนินการ' จาก status-filter-select",
   "- แสดงเฉพาะ order status=pending\n- ไม่เห็นสถานะอื่น",
   "status-filter-select, order-row","GET /api/admin/orders?status=pending","","To Be Automated","tests/admin/orders.spec.ts","admin,orders,filter"),

  ("TC-AORD-03","Admin Orders","Filter","Filter ไม่พบ order ในสถานะนั้น","P2","Functional","Admin",
   "ไม่มี order status=delivered",
   "1. เลือก 'ได้รับแล้ว'",
   "- Empty state 'ไม่มีออเดอร์'",
   "no-orders","GET /api/admin/orders?status=delivered","","To Be Automated","tests/admin/orders.spec.ts","admin,orders,filter,empty"),

  ("TC-AORD-04","Admin Orders","Update Status","อัปเดตสถานะ order","P0","Smoke","Admin",
   "มี order status=pending",
   "1. หา order status=pending\n2. เปลี่ยน dropdown เป็น 'confirmed'",
   "- Toast 'อัปเดตสถานะออเดอร์ #{id} แล้ว'\n- Badge เปลี่ยนเป็น 'ยืนยันแล้ว'",
   "status-select-{id}, order-status-{id}, toast","PUT /api/admin/orders/{id}/status","","To Be Automated","tests/admin/orders.spec.ts","admin,orders,status,smoke"),

  ("TC-AORD-05","Admin Orders","Update Status","อัปเดตสถานะเป็น cancelled","P1","Functional","Admin",
   "มี order ใดก็ได้",
   "1. เลือก 'cancelled' จาก status dropdown",
   "- Badge เป็น 'ยกเลิก' + red style\n- Order ยังอยู่ใน list (ไม่ลบ)",
   "status-select-{id}, order-status-{id}","PUT /api/admin/orders/{id}/status","","To Be Automated","tests/admin/orders.spec.ts","admin,orders,status,cancel"),

  ("TC-AORD-06","Admin Orders","Navigation","คลิก order link ไป detail","P2","UI","Admin",
   "—",
   "1. คลิก '#123' link ใน order row",
   "- Navigate ไป /order-detail.html?id=123\n- แสดงรายละเอียดครบ",
   "order-row","—","","To Be Automated","tests/admin/orders.spec.ts","admin,orders,navigation"),

  # ── I18N ──
  ("TC-I18N-01","i18n","Language Switch","เปลี่ยนเป็นภาษาอังกฤษ","P1","Functional","Guest",
   "ภาษาปัจจุบันเป็นไทย",
   "1. คลิก 🌐 ใน navbar\n2. คลิก 'English'",
   "- หน้า reload\n- nav links: Home, Products, Cart, Orders\n- CTA: 'Shop Now →'",
   "nav-user-menu","—","","To Be Automated","tests/i18n/i18n.spec.ts","i18n,language-switch"),

  ("TC-I18N-02","i18n","Language Switch","เปลี่ยนกลับเป็นภาษาไทย","P1","Functional","Guest",
   "ภาษาปัจจุบันเป็น English",
   "1. คลิก 🌐 → 'ภาษาไทย'",
   "- หน้า reload\n- ข้อความทั้งหมดกลับเป็นไทย",
   "","—","","To Be Automated","tests/i18n/i18n.spec.ts","i18n,language-switch"),

  ("TC-I18N-03","i18n","Persistence","ภาษาคงอยู่หลัง reload","P2","Functional","Guest",
   "เปลี่ยนเป็น English แล้ว",
   "1. กด F5 / reload",
   "- ยังเป็น English (localStorage คงอยู่)",
   "","—","","To Be Automated","tests/i18n/i18n.spec.ts","i18n,persistence"),

  ("TC-I18N-04","i18n","Dynamic Content","Dynamic content ใช้ภาษาที่เลือก","P2","Functional","User",
   "ภาษา English, Login แล้ว, มีออเดอร์",
   "1. เปิด /orders.html ขณะ English",
   "- Status badge แสดง 'Pending', 'Confirmed' (ไม่ใช่ภาษาไทย)",
   "order-status","—","","To Be Automated","tests/i18n/i18n.spec.ts","i18n,dynamic-content"),

  ("TC-I18N-05","i18n","Toast","Toast ใช้ภาษาที่เลือก","P2","Functional","User",
   "ภาษา English",
   "1. ลบ item ออกจาก cart",
   "- Toast แสดง 'Item removed' (ไม่ใช่ภาษาไทย)",
   "toast","—","","To Be Automated","tests/i18n/i18n.spec.ts","i18n,toast"),

  # ── NAVIGATION ──
  ("TC-NAV-01","Navigation","Guest Navbar","Navbar ของ Guest","P0","Smoke","Guest",
   "ไม่ได้ login",
   "1. เปิดหน้าใดก็ได้",
   "- เห็น: logo, Home, Products, Login, Register\n- ไม่เห็น: cart, orders, admin, user menu",
   "nav-logo, nav-home, nav-products, nav-login, nav-register","—","","To Be Automated","tests/navigation/nav.spec.ts","nav,smoke,guest"),

  ("TC-NAV-02","Navigation","User Navbar","Navbar ของ User","P0","Smoke","User",
   "Login ด้วย testuser",
   "1. Login แล้วดู navbar",
   "- เห็น: logo, Home, Products, Cart+badge, Orders, user menu\n- ไม่เห็น: Login, Register, Admin",
   "nav-cart, cart-count, nav-orders, nav-user-menu","—","","To Be Automated","tests/navigation/nav.spec.ts","nav,smoke,user"),

  ("TC-NAV-03","Navigation","Admin Navbar","Navbar ของ Admin","P1","Functional","Admin",
   "Login ด้วย admin",
   "1. Login แล้วดู navbar",
   "- เห็น nav-admin link เพิ่มเติม",
   "nav-admin","—","","To Be Automated","tests/navigation/nav.spec.ts","nav,admin"),

  ("TC-NAV-04","Navigation","Dropdown","User dropdown เปิด/ปิด","P2","UI","User",
   "Login แล้ว",
   "1. คลิก user menu button\n2. ดู dropdown แสดง\n3. คลิกอีกครั้ง\n4. ดู dropdown ซ่อน",
   "- Dropdown toggle ด้วย click\n- มี: email, Profile link, Logout",
   "nav-user-menu, nav-profile, nav-logout","—","","To Be Automated","tests/navigation/nav.spec.ts","nav,dropdown,ui"),

  ("TC-NAV-05","Navigation","Dropdown","Language dropdown เปิด/ปิด","P2","UI","Guest",
   "—",
   "1. คลิก 🌐\n2. ดู lang-menu แสดง\n3. คลิก outside\n4. ดู lang-menu ซ่อน",
   "- Toggle ด้วย click\n- Click outside ปิด\n- ไม่หายเมื่อเลื่อน cursor เข้า dropdown",
   "","—","","To Be Automated","tests/navigation/nav.spec.ts","nav,dropdown,i18n,ui"),

  ("TC-NAV-06","Navigation","Dropdown","เปิด menu แรก ปิดอีก menu","P2","UI","User",
   "—",
   "1. เปิด user dropdown\n2. คลิก 🌐",
   "- User dropdown ปิด\n- Language dropdown เปิด (ไม่เปิดพร้อมกัน)",
   "nav-user-menu","—","","To Be Automated","tests/navigation/nav.spec.ts","nav,dropdown,ui"),

  ("TC-NAV-07","Navigation","Mobile","Mobile menu toggle","P3","UI","Guest",
   "Viewport < 768px",
   "1. Resize viewport < 768px\n2. คลิก hamburger\n3. ดู mobile menu\n4. คลิกอีกครั้ง",
   "- Mobile menu แสดง/ซ่อน\n- มี Home, Products links",
   "","—","","To Be Automated","tests/navigation/nav.spec.ts","nav,mobile,ui"),
]

# ─── Test Data Sheet ──────────────────────────────────────────────────────────
TEST_DATA = [
  # id, tc_ref, scenario, username, email, password, confirm_pw, full_name, notes
  ("TD-AUTH-01","TC-AUTH-01","Register valid","newuser01","newuser01@test.com","pass1234","pass1234","Test New User","happy path"),
  ("TD-AUTH-02","TC-AUTH-02","Duplicate username","testuser","other@test.com","pass1234","pass1234","Other User","testuser already exists"),
  ("TD-AUTH-05","TC-AUTH-05","Login user","testuser","","test1234","","","seed account"),
  ("TD-AUTH-06","TC-AUTH-06","Login admin","admin","","admin1234","","","seed account"),
  ("TD-AUTH-07","TC-AUTH-07","Wrong password","testuser","","wrongpass","","","should fail"),
  ("TD-PROD-02","TC-PROD-02","Search Nike","","","","","","keyword=Nike, expect ≥4 results"),
  ("TD-DETAIL-04","TC-DETAIL-04","Add to cart","","","","","","product_id=1, size=42, qty=2"),
  ("TD-CHK-04","TC-CHK-04","Place order","","","","","","shipping_name=John Doe, city=Bangkok, postal=10110, phone=081-000-0000, payment=credit_card"),
  ("TD-PROF-02","TC-PROF-02","Update profile","","updated@test.com","","","Updated Name","restore after test"),
  ("TD-PROF-03","TC-PROF-03","Change password","","","test1234","newpass99","","restore: reset to test1234 after test"),
  ("TD-PROF-06","TC-PROF-06","Save address","","","","","","name=Test Shipper, city=Chiang Mai, postal=50000, payment=bank_transfer"),
  ("TD-APROD-02","TC-APROD-02","Add product","","","","","","name=Test Shoe X, brand=TestBrand, price=1990, stock=10, sizes=40,41,42"),
]

# ─── Environment Config ───────────────────────────────────────────────────────
ENVS = [
  ("DEV",     "http://localhost:8000",  "http://localhost:8000/api", "admin","admin1234","testuser","test1234","true","true", "รันจาก run.sh / run.bat"),
  ("QA",      "https://shoeshub-qa.onrender.com",     "https://shoeshub-qa.onrender.com/api",      "admin","admin1234","testuser","test1234","false","true",  "Auto-deploy จาก branch: qa"),
  ("STAGING", "https://shoeshub-staging.onrender.com","https://shoeshub-staging.onrender.com/api",  "admin","admin1234","testuser","test1234","false","false", "Auto-deploy จาก branch: staging"),
  ("PROD",    "https://shoeshub-prod.onrender.com",   "https://shoeshub-prod.onrender.com/api",     "admin","admin1234","testuser","test1234","false","false", "Auto-deploy จาก branch: main"),
]

# ─── Build Workbook ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Test Cases
# ════════════════════════════════════════════════════════════════════════════════
ws_tc = wb.active
ws_tc.title = "📋 Test Cases"
ws_tc.sheet_view.showGridLines = False
ws_tc.freeze_panes = "A3"

TC_COLS = [
    ("TC_ID",             15),
    ("Module",            16),
    ("Feature",           18),
    ("Test Name (TH)",    32),
    ("Priority",          10),
    ("Type",              16),
    ("Role",              10),
    ("Precondition",      26),
    ("Test Steps",        52),
    ("Expected Result",   40),
    ("Key data-testids",  34),
    ("API Endpoint",      26),
    ("Test Data Ref",     14),
    ("Automation Status", 18),
    ("Script Path",       38),
    ("Tags",              30),
]

# Title row
ws_tc.merge_cells("A1:P1")
t = ws_tc.cell(row=1, column=1, value="🛒  ShoesHub — Test Cases")
t.fill = fx("EAB308"); t.font = ft(True,"1F2937",14)
t.alignment = al("center","center",False)
ws_tc.row_dimensions[1].height = 32

# Header row
for ci, (name, w) in enumerate(TC_COLS, 1):
    hdr(ws_tc, 2, ci, name)
    ws_tc.column_dimensions[get_column_letter(ci)].width = w
ws_tc.row_dimensions[2].height = 30

# Data rows
for ri, tc in enumerate(TC, 3):
    (tc_id, module, feature, name_th, priority, type_, role,
     precondition, steps, expected, testids, api, data_ref,
     auto_status, script_path, tags) = tc

    bg_row = "F9FAFB" if ri % 2 == 0 else "FFFFFF"
    vals = [tc_id, module, feature, name_th, priority, type_, role,
            precondition, steps, expected, testids, api, data_ref,
            auto_status, script_path, tags]

    for ci, v in enumerate(vals, 1):
        c = cell(ws_tc, ri, ci, v, bg=bg_row)

    # Priority color
    p_bg, p_fc = PRIORITY_STYLE.get(priority, ("E5E7EB","1F2937"))
    pc = ws_tc.cell(row=ri, column=5)
    pc.fill = fx(p_bg); pc.font = ft(True, p_fc, 9)
    pc.alignment = al("center","center",False); pc.border = bd()

    # Automation status color
    a_bg, a_fc = AUTO_STYLE.get(auto_status, ("E5E7EB","1F2937"))
    ac = ws_tc.cell(row=ri, column=14)
    ac.fill = fx(a_bg); ac.font = ft(True, a_fc, 9)
    ac.alignment = al("center","center",False); ac.border = bd()

    ws_tc.row_dimensions[ri].height = 72

# Data validation — Priority
dv_p = DataValidation(type="list", formula1='"P0,P1,P2,P3"', showDropDown=False)
ws_tc.add_data_validation(dv_p); dv_p.sqref = f"E3:E{2+len(TC)}"

# Data validation — Automation Status
dv_a = DataValidation(type="list",
    formula1='"Automated,To Be Automated,In Progress,Manual Only"',
    showDropDown=False)
ws_tc.add_data_validation(dv_a); dv_a.sqref = f"N3:N{2+len(TC)}"

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Test Run Log
# ════════════════════════════════════════════════════════════════════════════════
ws_run = wb.create_sheet("▶ Test Run Log")
ws_run.sheet_view.showGridLines = False
ws_run.freeze_panes = "A3"

RUN_COLS = [
    ("Run ID",          12),("Run Date",     13),("Environment",  14),
    ("Branch",          12),("Tester",        15),("TC_ID",        15),
    ("Test Name",       32),("Status",        12),("Actual Result",35),
    ("Bug ID",          12),("Duration (ms)", 14),("Notes",        25),
]

ws_run.merge_cells("A1:L1")
t2 = ws_run.cell(row=1, column=1, value="▶  ShoesHub — Test Run Log")
t2.fill = fx("1F2937"); t2.font = ft(True,"FFFFFF",14)
t2.alignment = al("center","center",False)
ws_run.row_dimensions[1].height = 32

for ci, (name, w) in enumerate(RUN_COLS, 1):
    hdr(ws_run, 2, ci, name)
    ws_run.column_dimensions[get_column_letter(ci)].width = w
ws_run.row_dimensions[2].height = 28

STATUS_STYLE = {
    "Pass":    ("DCFCE7","166534"),
    "Fail":    ("FEE2E2","991B1B"),
    "Blocked": ("FED7AA","9A3412"),
    "Skip":    ("F3F4F6","6B7280"),
    "Not Run": ("F9FAFB","9CA3AF"),
}

# Pre-fill one run entry per test case (Not Run)
run_id = f"RUN-{TODAY}-001"
for ri, tc in enumerate(TC, 3):
    bg = "F9FAFB" if ri % 2 == 0 else "FFFFFF"
    vals_run = [run_id, TODAY, "QA", "qa", "", tc[0], tc[3],
                "Not Run", "", "", "", ""]
    for ci, v in enumerate(vals_run, 1):
        cell(ws_run, ri, ci, v, bg=bg)

    # Status color
    s_bg, s_fc = STATUS_STYLE["Not Run"]
    sc = ws_run.cell(row=ri, column=8)
    sc.fill = fx(s_bg); sc.font = ft(True, s_fc, 9)
    sc.alignment = al("center","center",False); sc.border = bd()
    ws_run.row_dimensions[ri].height = 20

dv_status = DataValidation(type="list",
    formula1='"Pass,Fail,Blocked,Skip,Not Run"', showDropDown=False)
ws_run.add_data_validation(dv_status)
dv_status.sqref = f"H3:H{2+len(TC)}"

dv_env_run = DataValidation(type="list",
    formula1='"DEV,QA,STAGING,PROD"', showDropDown=False)
ws_run.add_data_validation(dv_env_run)
dv_env_run.sqref = f"C3:C{2+len(TC)}"

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Test Data
# ════════════════════════════════════════════════════════════════════════════════
ws_td = wb.create_sheet("📁 Test Data")
ws_td.sheet_view.showGridLines = False
ws_td.freeze_panes = "A3"

TD_COLS = [
    ("Data Set ID",   14),("TC Reference",  14),("Scenario",     22),
    ("Username",      14),("Email",          24),("Password",     14),
    ("Confirm PW",    14),("Full Name",      20),("Notes",        35),
]

ws_td.merge_cells("A1:I1")
t3 = ws_td.cell(row=1, column=1, value="📁  ShoesHub — Test Data")
t3.fill = fx("7C3AED"); t3.font = ft(True,"FFFFFF",14)
t3.alignment = al("center","center",False)
ws_td.row_dimensions[1].height = 32

for ci, (name, w) in enumerate(TD_COLS, 1):
    hdr(ws_td, 2, ci, name, bg="5B21B6")
    ws_td.column_dimensions[get_column_letter(ci)].width = w
ws_td.row_dimensions[2].height = 28

for ri, td in enumerate(TEST_DATA, 3):
    bg = "FAF5FF" if ri % 2 == 0 else "FFFFFF"
    for ci, v in enumerate(td, 1):
        cell(ws_td, ri, ci, v, bg=bg)
    ws_td.row_dimensions[ri].height = 22

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Environments
# ════════════════════════════════════════════════════════════════════════════════
ws_env = wb.create_sheet("🌐 Environments")
ws_env.sheet_view.showGridLines = False

ENV_COLS = [
    ("Environment",  14),("Base URL",      42),("API URL",       42),
    ("Admin User",   13),("Admin Pass",    13),("Test User",     13),
    ("Test Pass",    12),("Debug",          8),("Auto Seed",      10),
    ("Notes",        32),
]

ws_env.merge_cells("A1:J1")
t4 = ws_env.cell(row=1, column=1, value="🌐  Environments")
t4.fill = fx("0369A1"); t4.font = ft(True,"FFFFFF",14)
t4.alignment = al("center","center",False)
ws_env.row_dimensions[1].height = 32

for ci, (name, w) in enumerate(ENV_COLS, 1):
    hdr(ws_env, 2, ci, name, bg="075985")
    ws_env.column_dimensions[get_column_letter(ci)].width = w
ws_env.row_dimensions[2].height = 28

ENV_BG = {"DEV":"ECFDF5","QA":"EFF6FF","STAGING":"FFF7ED","PROD":"FEF2F2"}
ENV_LABEL_BG = {"DEV":"16A34A","QA":"1D4ED8","STAGING":"EA580C","PROD":"DC2626"}

for ri, ev in enumerate(ENVS, 3):
    env_name = ev[0]
    bg = ENV_BG.get(env_name, "FFFFFF")
    for ci, v in enumerate(ev, 1):
        cell(ws_env, ri, ci, v, bg=bg)
    lb = ws_env.cell(row=ri, column=1)
    lb.fill = fx(ENV_LABEL_BG.get(env_name,"374151"))
    lb.font = ft(True,"FFFFFF",10); lb.alignment = al("center","center",False)
    ws_env.row_dimensions[ri].height = 26

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Bug Report
# ════════════════════════════════════════════════════════════════════════════════
ws_bug = wb.create_sheet("🐛 Bug Report")
ws_bug.sheet_view.showGridLines = False
ws_bug.freeze_panes = "A3"

BUG_COLS = [
    ("Bug ID",        12),("Found Date",   13),("TC_ID",         14),
    ("Environment",   13),("Severity",     12),("Priority",      10),
    ("Title",         38),("Steps to Repro",40),("Expected",      30),
    ("Actual",        30),("Status",        13),("Assignee",      14),
    ("Fixed Date",    13),("Notes",         25),
]

ws_bug.merge_cells("A1:N1")
t5 = ws_bug.cell(row=1, column=1, value="🐛  Bug Report")
t5.fill = fx("DC2626"); t5.font = ft(True,"FFFFFF",14)
t5.alignment = al("center","center",False)
ws_bug.row_dimensions[1].height = 32

for ci, (name, w) in enumerate(BUG_COLS, 1):
    hdr(ws_bug, 2, ci, name, bg="991B1B")
    ws_bug.column_dimensions[get_column_letter(ci)].width = w
ws_bug.row_dimensions[2].height = 28

dv_sev = DataValidation(type="list",
    formula1='"Critical,High,Medium,Low"', showDropDown=False)
ws_bug.add_data_validation(dv_sev); dv_sev.sqref = "E3:E500"

dv_bug_status = DataValidation(type="list",
    formula1='"Open,In Progress,Fixed,Verified,Closed,Won\'t Fix"',
    showDropDown=False)
ws_bug.add_data_validation(dv_bug_status); dv_bug_status.sqref = "K3:K500"

for ri in range(3, 13):
    bg = "FFF5F5" if ri % 2 == 0 else "FFFFFF"
    for ci in range(1, 15):
        cell(ws_bug, ri, ci, "", bg=bg)
    ws_bug.row_dimensions[ri].height = 22

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Summary
# ════════════════════════════════════════════════════════════════════════════════
ws_sum = wb.create_sheet("📊 Summary")
ws_sum.sheet_view.showGridLines = False

ws_sum.column_dimensions["A"].width = 22
ws_sum.column_dimensions["B"].width = 14
ws_sum.column_dimensions["C"].width = 14
ws_sum.column_dimensions["D"].width = 22
ws_sum.column_dimensions["E"].width = 14

ws_sum.merge_cells("A1:E1")
t6 = ws_sum.cell(row=1, column=1, value="📊  Test Summary")
t6.fill = fx("EAB308"); t6.font = ft(True,"1F2937",14)
t6.alignment = al("center","center",False)
ws_sum.row_dimensions[1].height = 32

# Count by module
from collections import Counter
module_counts = Counter(tc[1] for tc in TC)
priority_counts = Counter(tc[4] for tc in TC)
type_counts = Counter(tc[5] for tc in TC)
auto_counts = Counter(tc[13] for tc in TC)

def section_hdr(ws, row, col, text, bg="1F2937"):
    ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(col+1)}{row}")
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fx(bg); c.font = ft(True,"FFFFFF",10)
    c.alignment = al("center","center",False); c.border = bd()

def sum_row(ws, row, col, label, count, bg="FFFFFF", label_bg=None):
    c1 = ws.cell(row=row, column=col, value=label)
    c1.fill = fx(label_bg or bg); c1.font = ft(False,"1F2937",9)
    c1.alignment = al("left","center",False); c1.border = bd()
    c2 = ws.cell(row=row, column=col+1, value=count)
    c2.fill = fx(bg); c2.font = ft(True,"1F2937",10)
    c2.alignment = al("center","center",False); c2.border = bd()
    ws.row_dimensions[row].height = 20

# Total
ws_sum.merge_cells("A3:E3")
total_cell = ws_sum.cell(row=3, column=1, value=f"Total Test Cases: {len(TC)}")
total_cell.fill = fx("F9FAFB"); total_cell.font = ft(True,"1F2937",12)
total_cell.alignment = al("center","center",False); total_cell.border = bd()
ws_sum.row_dimensions[3].height = 28

# By Module (col A-B)
section_hdr(ws_sum, 5, 1, "By Module")
for i, (mod, cnt) in enumerate(sorted(module_counts.items()), 6):
    sum_row(ws_sum, i, 1, mod, cnt, "FFFFFF" if i%2==0 else "F9FAFB")

# By Priority (col D-E)
section_hdr(ws_sum, 5, 4, "By Priority")
p_colors = {"P0":"FEE2E2","P1":"FED7AA","P2":"FEF9C3","P3":"DCFCE7"}
for i, (pri, cnt) in enumerate(sorted(priority_counts.items()), 6):
    sum_row(ws_sum, i, 4, f"{pri} — {'Smoke' if pri=='P0' else 'High' if pri=='P1' else 'Medium' if pri=='P2' else 'Low'}", cnt, p_colors.get(pri,"F9FAFB"))

# By Type
type_start = 6 + max(len(module_counts), len(priority_counts)) + 2
section_hdr(ws_sum, type_start, 1, "By Type")
for i, (t_, cnt) in enumerate(sorted(type_counts.items()), type_start+1):
    sum_row(ws_sum, i, 1, t_, cnt, "FFFFFF" if i%2==0 else "F9FAFB")

# By Automation Status
section_hdr(ws_sum, type_start, 4, "Automation Status")
a_colors = {"Automated":"DBEAFE","To Be Automated":"EDE9FE","In Progress":"DBEAFE","Manual Only":"F3F4F6"}
for i, (a_, cnt) in enumerate(sorted(auto_counts.items()), type_start+1):
    sum_row(ws_sum, i, 4, a_, cnt, a_colors.get(a_,"F9FAFB"))

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 7 — Legend / README
# ════════════════════════════════════════════════════════════════════════════════
ws_leg = wb.create_sheet("📖 Legend")
ws_leg.sheet_view.showGridLines = False
ws_leg.column_dimensions["A"].width = 22
ws_leg.column_dimensions["B"].width = 45
ws_leg.column_dimensions["D"].width = 22
ws_leg.column_dimensions["E"].width = 45

ws_leg.merge_cells("A1:E1")
tl = ws_leg.cell(row=1, column=1, value="📖  Legend & How To Use")
tl.fill = fx("374151"); tl.font = ft(True,"FFFFFF",14)
tl.alignment = al("center","center",False)
ws_leg.row_dimensions[1].height = 32

legend_data = [
    ("PRIORITY", [
        ("P0 — Smoke/Critical","Must pass before release. Run on every deploy.",       "DC2626","FFFFFF"),
        ("P1 — High",          "Core user journeys. Run on every regression.",          "EA580C","FFFFFF"),
        ("P2 — Medium",        "Important features. Run on full regression.",           "CA8A04","FFFFFF"),
        ("P3 — Low",           "Edge cases / UI polish. Run on release regression.",    "16A34A","FFFFFF"),
    ]),
    ("AUTOMATION STATUS", [
        ("Automated",          "Script exists and runs in CI.",                         "1D4ED8","FFFFFF"),
        ("To Be Automated",    "Planned for automation. Script not written yet.",       "7C3AED","FFFFFF"),
        ("In Progress",        "Script being written.",                                 "0369A1","FFFFFF"),
        ("Manual Only",        "Not suitable for automation (e.g., visual/backdrop).", "6B7280","FFFFFF"),
    ]),
    ("TEST RUN STATUS", [
        ("Pass",    "Test passed as expected.",                "166534","DCFCE7"),
        ("Fail",    "Test failed. Log bug in Bug Report.",    "991B1B","FEE2E2"),
        ("Blocked", "Cannot run due to dependency/env issue.","9A3412","FED7AA"),
        ("Skip",    "Intentionally skipped this run.",        "6B7280","F3F4F6"),
        ("Not Run", "Not yet executed.",                      "9CA3AF","F9FAFB"),
    ]),
    ("ENVIRONMENTS", [
        ("DEV",     "Local machine. run.sh / run.bat. AUTO_SEED=true.",   "FFFFFF","16A34A"),
        ("QA",      "Render branch: qa. AUTO_SEED=true (fresh per deploy).","FFFFFF","1D4ED8"),
        ("STAGING", "Render branch: staging. AUTO_SEED=false.",           "FFFFFF","EA580C"),
        ("PROD",    "Render branch: main. Production. AUTO_SEED=false.",  "FFFFFF","DC2626"),
    ]),
]

how_to = [
    "HOW TO USE THIS FILE",
    "",
    "1. Test Cases sheet — ข้อมูล test case ทั้งหมด 78 รายการ",
    "   ▸ Priority: P0=Smoke, P1=High, P2=Medium, P3=Low",
    "   ▸ Tags column ใช้สำหรับ Playwright --grep flag",
    "   ▸ Script Path ระบุตำแหน่ง spec file ใน Playwright project",
    "",
    "2. Test Run Log — บันทึกผลการทดสอบแต่ละ run",
    "   ▸ สร้าง Run ID ใหม่ทุกครั้งที่ run (เช่น RUN-2026-04-22-001)",
    "   ▸ อัปเดต Status: Pass/Fail/Blocked/Skip/Not Run",
    "   ▸ ถ้า Fail → บันทึก Bug ID ใน column Bug ID",
    "",
    "3. Test Data — ข้อมูลสำหรับ data-driven testing",
    "   ▸ Playwright อ่าน sheet นี้ผ่าน xlsx library",
    "   ▸ กรอง Auto Status = 'Automated' แล้ว map TD_ ref",
    "",
    "4. Bug Report — บันทึก bug ที่พบระหว่างทดสอบ",
    "   ▸ Bug ID format: BUG-YYYYMMDD-###",
    "   ▸ Severity: Critical/High/Medium/Low",
    "",
    "5. Automation Linkage (Playwright)",
    "   const wb = xlsx.readFile('ShoesHub_Test_Cases.xlsx')",
    "   const tcs = xlsx.utils.sheet_to_json(wb.Sheets['📋 Test Cases'])",
    "   const toRun = tcs.filter(t => t.Automation_Status === 'Automated')",
    "   // use t.Tags with --grep, t.TC_ID as test annotation",
]

row = 3
for item in how_to:
    c = ws_leg.cell(row=row, column=4, value=item)
    if item.startswith("HOW TO"):
        ws_leg.merge_cells(f"D{row}:E{row}")
        c.fill = fx("1F2937"); c.font = ft(True,"FFFFFF",10)
        c.alignment = al("left","center",False)
    elif item.startswith("   ▸"):
        c.font = ft(False,"6B7280",9,True); c.alignment = al("left","top",True)
    elif item and not item.startswith(" "):
        c.font = ft(True,"374151",9); c.alignment = al("left","top",False)
    else:
        c.font = ft(False,"374151",9); c.alignment = al("left","top",True)
    c.border = bd("E5E7EB")
    ws_leg.row_dimensions[row].height = 16
    row += 1

row = 3
for section_name, items in legend_data:
    section_hdr(ws_leg, row, 1, section_name)
    row += 1
    for label, desc, fc, bg in items:
        c1 = ws_leg.cell(row=row, column=1, value=label)
        c1.fill = fx(bg); c1.font = ft(True, fc, 9)
        c1.alignment = al("center","center",False); c1.border = bd()
        c2 = ws_leg.cell(row=row, column=2, value=desc)
        c2.font = ft(False,"374151",9); c2.alignment = al("left","center",True)
        c2.border = bd()
        ws_leg.row_dimensions[row].height = 20
        row += 1
    row += 1

# ─── Tab Colors ───────────────────────────────────────────────────────────────
ws_tc.sheet_properties.tabColor  = "EAB308"
ws_run.sheet_properties.tabColor = "22C55E"
ws_td.sheet_properties.tabColor  = "7C3AED"
ws_env.sheet_properties.tabColor = "0369A1"
ws_bug.sheet_properties.tabColor = "DC2626"
ws_sum.sheet_properties.tabColor = "374151"
ws_leg.sheet_properties.tabColor = "6B7280"

# ─── Save ─────────────────────────────────────────────────────────────────────
out = "ShoesHub_Test_Cases.xlsx"
wb.save(out)
print(f"[OK] Created: {out}")
print(f"     Sheets : 7 (Test Cases, Test Run Log, Test Data, Environments, Bug Report, Summary, Legend)")
print(f"     TCs    : {len(TC)} test cases across {len(set(tc[1] for tc in TC))} modules")
